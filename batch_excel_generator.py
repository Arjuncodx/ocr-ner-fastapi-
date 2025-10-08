#!/usr/bin/env python
"""
batch_excel_generator.py - Generate Master Excel for Batch Processing
====================================================================
Creates single Excel file with all batch results in rows.

Senior Python OCR Developer - Fortune 500 Grade
October 2025
"""

import logging
from pathlib import Path
from typing import List, Dict
from datetime import datetime

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("⚠️  openpyxl not available - batch Excel generation disabled")


# ============================================================================
# BATCH EXCEL GENERATOR
# ============================================================================
class BatchExcelGenerator:
    """Generate master Excel file with all batch documents"""
    
    def __init__(self):
        """Initialize batch Excel generator"""
        self.styles = self._init_styles()
        logger.info("✅ BatchExcelGenerator initialized")
    
    def _init_styles(self):
        """Initialize Excel styles"""
        if not OPENPYXL_AVAILABLE:
            return {}
        
        return {
            'header': {
                'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'data': {
                'font': Font(name='Calibri', size=11),
                'alignment': Alignment(horizontal='left', vertical='center', wrap_text=False),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'title': {
                'font': Font(name='Calibri', size=16, bold=True, color='1F4E78'),
                'alignment': Alignment(horizontal='center', vertical='center')
            }
        }
    
    def generate_master_excel(
        self,
        extracted_data: List[Dict[str, str]],
        output_path: Path,
        template_name: str = "Batch Processing",
        batch_id: str = ""
    ) -> bool:
        """
        Generate master Excel with all documents as rows
        
        Args:
            extracted_data: List of dictionaries (one per document)
            output_path: Where to save the Excel file
            template_name: Name of template used
            batch_id: Batch processing ID
            
        Returns:
            bool: Success status
        """
        try:
            if not OPENPYXL_AVAILABLE:
                logger.warning("⚠️  OpenPyXL not available")
                return False
            
            if not extracted_data:
                logger.warning("⚠️  No data to generate Excel")
                return False
            
            logger.info(f"\n📊 GENERATING MASTER BATCH EXCEL")
            logger.info(f"   Documents: {len(extracted_data)}")
            
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # Create main data sheet
            ws = wb.create_sheet("Batch Results", 0)
            
            # Title
            title_text = f"OCR Elite v15.2 - Batch Processing Results - {template_name}"
            ws['A1'] = title_text
            ws['A1'].font = self.styles['title']['font']
            ws['A1'].alignment = self.styles['title']['alignment']
            
            # Get all column names (from first document)
            columns = list(extracted_data[0].keys())
            
            # Remove internal fields (starting with _)
            display_columns = [col for col in columns if not col.startswith('_')]
            
            # Merge title cells
            merge_end = get_column_letter(len(display_columns))
            ws.merge_cells(f'A1:{merge_end}1')
            ws.row_dimensions[1].height = 30
            
            # Info row
            ws.append([])
            ws.append([f"Batch ID: {batch_id}", f"Documents: {len(extracted_data)}", f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            
            # Empty row
            ws.append([])
            
            # Headers
            ws.append(display_columns)
            header_row = ws.max_row
            
            for col_idx, column_name in enumerate(display_columns, start=1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.font = self.styles['header']['font']
                cell.fill = self.styles['header']['fill']
                cell.alignment = self.styles['header']['alignment']
                cell.border = self.styles['header']['border']
                
                # Auto-width
                col_width = max(15, len(str(column_name)) + 2)
                ws.column_dimensions[get_column_letter(col_idx)].width = col_width
            
            # Data rows
            for doc_data in extracted_data:
                row_data = [doc_data.get(col, "") for col in display_columns]
                ws.append(row_data)
                
                # Apply styling to last row
                last_row = ws.max_row
                for col_idx in range(1, len(display_columns) + 1):
                    cell = ws.cell(row=last_row, column=col_idx)
                    cell.font = self.styles['data']['font']
                    cell.alignment = self.styles['data']['alignment']
                    cell.border = self.styles['data']['border']
            
            # Freeze panes
            ws.freeze_panes = f'A{header_row + 1}'
            
            # Save
            wb.save(output_path)
            logger.info(f"✅ Master Excel saved: {output_path.name}")
            return True
        
        except Exception as e:
            logger.error(f"❌ Batch Excel generation error: {e}")
            return False


# ============================================================================
# GLOBAL INSTANCE
# ============================================================================
batch_excel_generator = BatchExcelGenerator()
